import datetime
import threading
import openpyxl


class openpyxl_excel(object):
    _instance_lock = threading.Lock()
    
    def __init__(self, filename=None, index=0):
        if os.path.exists(filename):
            self.inwb = openpyxl.load_workbook(filename)  # 读文件
            self.sheetnames = self.inwb.get_sheet_names()  # 获取文件所以sheet
            self.ws = self.inwb.get_sheet_by_name(self.sheetnames[index])  # 默认获取第一个
            self.rows = self.ws.max_row
            self.cols = self.ws.max_column
            self.filename = filename
    
    def __new__(cls, *args, **kwargs):  # 单例
        if not hasattr(openpyxl_excel, "_instance"):
            with openpyxl_excel._instance_lock:
                if not hasattr(openpyxl_excel, "_instance"):
                    openpyxl_excel._instance = object.__new__(cls)
        return openpyxl_excel._instance
    
    def read_execl_by_cols(self, cols=1):  # 通过列分组获取
        if not self.inwb:
            return "初始化失败或文件不存在"
        if cols > self.cols:
            cols = self.cols
        rowss = self.ws.iter_cols(min_row=1, min_col=1, max_col=cols, max_row=self.rows)
        dict_cols = dict()
        i = 1
        for row in rowss:
            list = []
            for cell in row:
                list.append(cell.value)
            dict_cols[i] = list.copy()
            i += 1
        return dict_cols
    
    def read_execl_by_rows(self, rows=1):  # 通过行分组获取
        if not self.inwb:
            return "初始化失败或文件不存在"
        if rows > self.rows:
            rows = self.rows
        rowss = self.ws.iter_rows(min_row=1, min_col=1, max_col=self.cols, max_row=rows)
        dict_rows = dict()
        i = 1
        for row in rowss:
            list = []
            for cell in row:
                list.append(cell.value)
            dict_rows[i] = list.copy()
            i += 1
        return dict_rows
    
    def read_execl_all(self):  # 获取表格所以内容
        if not self.ws:
            return "初始化失败或文件不存在"
        dict_all = dict()
        i = 1
        for eow in self.ws.values:
            p_list = []
            for values in eow:
                p_list.append(values)
            dict_all[i] = p_list
            i += 1
        return dict_all
    
    def read_by_rows_cols(self, rows, cols):
        if not self.ws:
            return "初始化失败或文件不存在"
        return self.ws.cell(rows, cols).value
    
    @classmethod
    def write_new_file_excel(cls, savefilename="D:\\work\\Excel_txtProcesss\\test.csv", datadict=None):
        """写表格，datadict以每行存list或元组，总的以字典格式传递
        :param savefilename : 文件路径及名字
        :type savefilename : str
        :param datadict : 数据字典
        :type datadict : dict
        """
        if not datadict:
            return "空数据"
        try:
            outwb = openpyxl.Workbook()  # 打开一个将写的文件
            outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
        except Exception as e:
            print("Excel新建表：", e)
            return "处理新建表格失败"
        rows = 1
        for row in datadict:
            cols = 1
            datalist = datadict[row]
            for col in datalist:
                try:
                    outws.cell(rows, cols).value = col  # 写文件
                except Exception as e:
                    print(e, row, col)
                cols += 1
            rows += 1
        try:
            outwb.save(savefilename)  # 一定要记得保存
        except Exception as e:
            print("Excel保存：", e)
            return "文件保存失败"
        return "OK"
    
    def write_alter_all_excel(self, datadict):
        """ 覆盖式修改
        :param datadict: 数据字典
        :type datadict: dict
        :return: 修改结果
        :rtype: str
        """
        if not self.ws:
            return "初始化失败或文件不存在"
        if not datadict:
            return "空数据"
        rows = 1
        for row in datadict:
            cols = 1
            if datadict[row] is not list:
                return "数据格式请以list传递"
            for col in datadict[row]:
                self.ws.cell(rows, col).value = col  # 写文件
                cols += 1
            rows += 1
        try:
            self.inwb.save(self.filename)  # 一定要记得保存
        except Exception as e:
            print("Excel保存：", e)
            return "文件保存失败"
        return "OK"
    
    def write_all_csv(self, savefilename="D:\\work\\Excel_txtProcesss\\test.csv", datadict=None):
        """ 写csv文件
        :param savefilename : 文件路径及名字
        :type savefilename : str
        :param datadict : 数据字典
        :type datadict : dict
        :return: 操作结果
        :rtype: str
        """
        if not datadict:
            return "空数据"
        rows = 1
        name = ''
        for row in datadict:
            cols = 1
            l = ''
            if datadict[row] is not list:
                return "数据格式请以list传递"
            for col in datadict[row]:
                l += col + ','  # 写文件
                cols += 1
            rows += 1
            name += l[:-1] + '\n'
        try:
            with open(savefilename, mode='a+') as f:
                f.write(name)  # 一定要记得保存
                f.close()
        except Exception as e:
            print("Excel保存：", e)
            return "文件保存失败"
        return "OK"


if __name__ == '__main__':
    import gc
    import psutil
    import os
    
    # info = psutil.virtual_memory()
    # print(u'内存使用：', psutil.Process(os.getpid()).memory_info().rss)
    # print(u'总内存：', info.total)
    # print(u'内存占比：', info.percent)
    # print(u'cpu个数：', psutil.cpu_count())
    # print(gc.get_threshold())
    # print('执行前信息')
    # datadict = dict(t1=['a', 'v', 'b', 'd'], t2=['a', 'v', 'b', 'd', '3'])
    """ 拿数图片名称生成对应的数据
    datadict = dict()
    from axf.dbmysql import my_si_db
    from app.src.randoms_idcard import *
    list2 = my_si_db('select Num from face_putic', 'test')
    all = len(list2)
    a = 1
    for i in list2:
        try:
            i = int(i[0])
        except:
            i = i[0]
        if type(i) is int:
            IDcard = iD_card_run()
            datadict[i] = [i, 'test_' + str(i), '男', '新生', 0, '学生', '中国', '汉族', '身份证', IDcard, '中国广州', '2025/01/01 0:00:00', '正常', 13800000000 + i, 2]
            print('当前处理执行{0}个，总数{1}'.format(a, all))
            a += 1
        else:
            print('当前ID格式不符：', i)
            a += 1
    print('开始写Excel')
    aa = openpyxl_excel.write_new_file_excel("I:\\文档\\sn.csv", datadict)
    print(aa)
    """
    
    """
    # 个人档案 20200000000开始
    from app.src.randoms_idcard import *

    datadict = dict()
    a = 1
    all = 600000
    datadict[1] = ['学工号', '姓名' ,'性别' ,'部门' ,'宿舍ID' ,'身份' ,
                   '国籍' ,'民族' ,'证件类型', '证件号', '住址', '过期时间', '档案状态', '手机号', '职称',
                   '职级',
                   '学位', '学籍编号', '学籍所在地', '电子邮箱', '办公地点', '邮政编码', '家长联系电话', '政治面貌', '籍贯', '户口所在地', '婚姻状况', '是否是独生子女' ,
                   '出生时间' ,'卡号' ,'卡片状态' ,'处理方式']
    nu = 0
    s = 0
    for i in range(1, all + 1):
        xingbie = '男'
        if i % 2 == 0:
            xingbie = '女'
        IDcard = iD_card_run()
        datadict[i + 1 - nu] = [20200000000 + i, 'test_' + str(20200000000 + i), xingbie, '新生', 1, '学生', '中国', '汉族', '身份证', IDcard, '中国广州', '2025/01/01 0:00:00',
                       '正常', 13810000000 + i, '', '', '', '', '', '', '', '', '', '', '', '',
                           '未婚', '是', '', '', '', 2]
        print('当前处理执行{0}个，总数{1}'.format(a, all))
        a += 1
        if i % 200000 == 0:
            s += 1
            print('开始写Excel第' + str(s) + '的200000个')
            aa = openpyxl_excel.write_all_csv("I:\\文档\\人员-20万-{}.csv".format(s), datadict)
            print(aa)
            if aa == 'OK':
                datadict.clear()
                datadict[1] = ['学工号', '姓名', '性别', '部门', '宿舍ID', '身份', '国籍', '民族', '证件类型', '证件号', '住址', '过期时间', '档案状态',
                               '手机号', '职称', '职级', '学位', '学籍编号', '学籍所在地', '电子邮箱', '办公地点', '邮政编码', '家长联系电话', '政治面貌', '籍贯',
                               '户口所在地', '婚姻状况', '是否是独生子女', '出生时间', '卡号', '卡片状态', '处理方式']
                nu += 50000
            else:
                print('失败了', s, aa)
        """
    
    """  #  多层级部门
    # 导入部门
    datadict = dict()
    a = 1
    all = 80
    datadict[1] = ['部门']
    nu = 0
    s = 0
    name = ''
    for i in range(1, all + 1):
        if not name:
            name = "学院" + str(i)
            datadict[i + 1 - nu] = [name]
        else:
            name = name + '#' + '学院' + str(i)
            datadict[i + 1 - nu] = [name]
        print('当前处理执行{0}个，总数{1}'.format(a, all))
        a += 1
        if i % all == 0:
            s += 1
            print('开始写Excel第' + str(s) + '的20个')
            aa = openpyxl_excel.write_all_csv("I:\\文档\\bumen-{}.csv".format(s), datadict)
            print(aa)
            if aa == 'OK':
                datadict.clear()
                datadict[1] = ['部门']
            else:
                print('失败了', s, aa)
    """
    """
    # 多部门
    # 导入部门
    datadict = dict()
    a = 1
    all = 10001
    datadict[1] = ['部门']
    nu = 0
    s = 0
    name = ''
    for i in range(1, all + 1):
        if not name:
            name = "人名大学" + str(i)
            datadict[i + 1 - nu] = [name]
        else:
            dname = name + '#' + '人名大学' + str(i)
            datadict[i + 1 - nu] = [dname]
        print('当前处理执行{0}个，总数{1}'.format(a, all))
        a += 1
        if i % all == 0:
            s += 1
            print('开始写Excel第' + str(s) + '的80个')
            aa = openpyxl_excel.write_all_csv("I:\\文档\\bumen多级-{}.csv".format(s), datadict)
            print(aa)
            if aa == 'OK':
                datadict.clear()
                datadict[1] = ['部门']
            else:
                print('失败了', s, aa)
    """
    
    """
    # 部门变更
    datedict = dict()
    from axf.dbmysql import my_si_db
    s = 0
    name = ''
    a = 1
    all = 80
    res = my_si_db('select ID,UserNum from pub_userinfo LIMIT 1, {};'.format(all), 'identityface')
    datedict[1] = ['人员编号', '变更部门', '是否需要添加']
    for i, da in zip(range(1, all + 1), res):
        datedict[i + 1] = [str(da[1]), '学前班', 2]
        print('当前处理执行{0}个，总数{1}'.format(a, all))
        a += 1
        if i % all == 0:
            s += 1
            print('开始写Excel第' + str(s) + '的' + str(all) + '个')
            aa = openpyxl_excel.write_all_csv("I:\\文档\\部门变更-{}.csv".format(s), datedict)
            print(aa)
            if aa == 'OK':
                datedict.clear()
                datedict[1] = ['人员编号', '变更部门', '是否需要添加']
            else:
                print('失败了', s, aa)
    """

    """
    # 身份变更
    datedict = dict()
    from axf.dbmysql import my_si_db
    s = 0
    name = ''
    a = 1
    all = 20
    res = my_si_db('select ID,UserNum from pub_userinfo LIMIT 1, {};'.format(all), 'identityface')
    datedict[1] = ['人员编号', '变更身份编号']
    for i, da in zip(range(1, all + 1), res):
        datedict[i + 1] = [str(da[1]), 'XS0021']
        print('当前处理执行{0}个，总数{1}'.format(a, all))
        a += 1
        if i % all == 0:
            s += 1
            print('开始写Excel第' + str(s) + '的' + str(all) + '个')
            aa = openpyxl_excel.write_all_csv("I:\\文档\\身份变更-{}.csv".format(s), datedict)
            print(aa)
            if aa == 'OK':
                datedict.clear()
                datedict[1] = ['人员编号', '变更部门', '是否需要添加']
            else:
                print('失败了', s, aa)
    """
    
    """"""
    # 宿舍导入
    sushedate = dict()
    s = 0
    name = ''
    a = 1
    all = 50000
    nu = 0
    lao = ['品园', '知行', '北园', '红园', '宜园', '东风', '培训', 'test1', 'test2', 'test3']
    m = 6
    if (all / 6 / len(lao)) > 99:
        while True:
            m += 1
            new = int(all / m / len(lao))
            if new <= 99:
                break
    else:
        new = int(all / 6 / (len(lao) + 1))
    Num = 1
    name += '宿舍\n'  # 表头
    for l in lao:
        first = ''
        for i in range(1, m + 1):
            if not first:
                first = l + str(i) + '栋'
                name += first + '\n'
                for n in range(1, 100):
                    name += first + '#' + str(i) + '-' + str(n) + '\n'
                    print('当前处理执行{0}个，总数{1}'.format(a, all))
                    a += 1
                    if a > all:
                        break
            else:
                for n in range(1, 100):
                    name += first + '#' + str(i) + '-' + str(n) + '\n'
                    print('当前处理执行{0}个，总数{1}'.format(a, all))
                    a += 1
                    if a > all:
                        break
            if a > all:
                break
    print('开始写csv的' + str(all) + '个')
    with open('I:\\文档\\宿舍导入.csv', mode='a+') as f:
        f.write(name)
        f.close()
    
    # def task(arg):
    #     filename = r'I:\work\jiekou\c\10.csv'
    #     obj = openpyxl_excel(filename=filename)
    #     # print(obj.read_exel_all())
    #  # for i in range(10):
    #     t = threading.Thread(target=task, args=[i, ])
    #     t.start()  # print(gc.get_threshold())
    # print('回收内存完成')
    # print(u'内存使用：', psutil.Process(os.getpid()).memory_info().rss)
    # print(u'总内存：', info.total)
    # print(u'内存占比：', info.percent)
    # print(u'cpu个数：', psutil.cpu_count())
    # gc.collect()  # 回收内存
    # print('回收内存完成信息')
    # print(u'内存使用：', psutil.Process(os.getpid()).memory_info().rss)
    # print(u'总内存：', info.total)
    # print(u'内存占比：', info.percent)
    # print(u'cpu个数：', psutil.cpu_count())
    # print(gc.get_threshold())
